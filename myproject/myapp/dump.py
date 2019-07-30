import filesync
from openpyxl import load_workbook
from openpyxl import Workbook
import nlp
import os
def dump():
    l = filesync.syncfile()
    wb = Workbook()
    ws=wb.active
    rng=1
    for i in l:
        wb2=load_workbook(i)
        ws2=wb2.active
        row=2
        col=1
        #counting total number of entries in all KEDB files
        while True:
            cr=ws2.cell(row=row,column=col).value
            if cr is None:
                break
            else:
                rng+=1
            row+=1
    #creating master KEDB files with total no of rows counted and 6 columns
    for i in range(1, rng):
        for j in range(1, 7):
            ws.cell(row=i, column=j)
    m=1

    for i in l:
        row_no=2
        wb1=load_workbook(i)
        ws1=wb1.active
        while True:
            c=ws1.cell(row=row_no,column=1).value
            c1=ws1.cell(row=row_no,column=2).value
            c2=i
            text4 = c2
            inc = text4.split("\\")
            inc = inc[-1]
            leng1 = len(inc)
            inc1 = inc[0:leng1 - 5:]
            #print inc1


            c3=ws1.cell(row=row_no,column=3).value
            if c3 is None:
                #if no prerequisite then adding "not applicable in prerequisite section"
                c3="Not Applicable"
            if c is None:
                break
            else:
                ws.cell(row=m,column=1).value=c
                ws.cell(row=m,column=2).value=c1
                ws.cell(row=m,column=3).value=c2
                ws.cell(row=m,column=5).value=c3
                ws.cell(row=m, column=6).value = inc1
                m+=1
            row_no+=1
    #create KEDB.txt with only the known error column required for spell check
    fp=open("c:\\Python27\\myproject\\myapp\\KEDB.txt",'w')
    r=1
    while True:
        c = ws.cell(row=r, column=1).value
        if c is None:
            break
        else:
            c=c+'\n'
            fp.write(c)
        r += 1
    fp.close()
    r=1
    while True:
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=6).value
        # print type('c2')
        c=str(c1)
        tkn=nlp.create_keyword(c)
        #print type('tkn')
        # tkn=tkn+" "+c2
        if c1 is None:
            break
        else:
            tkn1=tkn
            c2 = c2.lower()
            tkn1 = tkn1 + " " + c2

            if len(tkn1)==0:
                tkn1 = "X"
            #print tkn1
            #print len(tkn1)
            ws.cell(row=r, column=4).value=tkn1
        r += 1
    for i in l:
        tm=os.path.getmtime("c:\\Python27\\myproject\\myapp\\MASTER.xlsx")
        if tm < os.path.getmtime(i):
            wb.save("c:\\Python27\\myproject\\myapp\\MASTER.xlsx")
            break

dump()