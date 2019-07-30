import wordnet
from openpyxl import load_workbook
import dump

global l8
l8=[]
def search(s):
    global l5,l6,l7,ltt
    l=wordnet.tokens(s)
    wb = load_workbook(filename='c:\\Python27\\myproject\\myapp\\MASTER.xlsx')
    ws=wb.active
    r=1
    l2=[]
    l3=[]
    l11=[]
    l7=[]
    l8=[]
    ltt=[]
    lt=[]
    while True:
        c = ws.cell(row=r, column=4).value
        c1=ws.cell(row=r,column=2).value
        c2=ws.cell(row=r,column=3).value
        c3=ws.cell(row=r,column=1).value
        c4=ws.cell(row=r,column=5).value

        if c is None:
            break
        else:
            l8.append(c3)
            l1 = c.split(' ')
            count=1
            count1=1
            #counting number of tokens matched
            for i in l:
                for j in l1:
                    count1+=1
                    if i == j:
                        count += 1
            #ratio of total comparison and match
            count2=count1/count
            l2.append(count2)
            l3.append(c1)
            l11.append(c2)
            lt.append(c4)
        r += 1
    #storing details if ratio is less than 500 in ranking of best match
    l4=[x for (x, y, z,z1) in sorted(zip(l2, l3, l11, lt)) if x<500 ]
    l5=[y for (x,y,z,z1) in sorted(zip(l2,l3,l11,lt)) if x<500 ]
    l6=[z for (x,y,z,z1) in sorted(zip(l2,l3,l11, lt)) if x<500]
    ltt=[z1 for (x,y,z,z1) in sorted(zip(l2,l3,l11,lt)) if x<500 ]
    for i in l6:
        k=i.split("\\")
        l7.append(k[-1])
#returning all the filenames
def get_file_name():
    fl=[]
    for i in l7:
        leng=len(i)
        i1=i[0:leng-5:]
        fl.append(i1)
    return fl
#reutrning all the solutions from the search
def get_search_result(s):
    #dump is created
    dump.dump()
    #searching is done
    search(s)
    return l5
#returning prerequisites
def pre():
    return ltt