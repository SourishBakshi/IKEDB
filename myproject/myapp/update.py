import openpyxl
from openpyxl.styles import Alignment
def update(question,answer,prerequisite,file_choice):
    fp=open("C:\\Python27\\myproject\\myapp\\access_to_kedbfiles.txt",'r')
    data=fp.read()
    f="{}\\{}.xlsx".format(data,file_choice)
    wb = openpyxl.load_workbook(f)
    ws=wb.active
    row_no=1
    print answer
    while True:
        c=ws.cell(row=row_no,column=1).value
        if c is None:
            ws.cell(row=row_no, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_no,column=1).value=question
            ws.cell(row=row_no,column=2).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_no,column=2).value=(answer)
            ws.cell(row=row_no, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_no,column=3).value=prerequisite
            break

        row_no+=1
    wb.save(f)

